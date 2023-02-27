import nltk

# nltk.download('averaged_perceptron_tagger')  # download the tagger

sent = ['The', 'horse', 'that', 'was', 'raced', 'past', 'the', 'barn', 'fell', 'down', '.']


print(nltk.pos_tag(sent)[4][1])


from openpyxl import Workbook
new_book = Workbook()
new_sheet = new_book.active
new_sheet['A1'] = 'Giuliana'
new_book.save("The_name_of_my_friend.xlsx")
sheet.max_row
# same
sheet_one.cell(row=3, column=1).value = 'La Habana'
sheet_one['A3'] = 'La Habana'